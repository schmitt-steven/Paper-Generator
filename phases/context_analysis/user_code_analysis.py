import os
import re
import json
import ast
import textwrap
from pathlib import Path
from dataclasses import dataclass, field
from pydantic import BaseModel
from settings import Settings
from utils.lazy_model_loader import LazyModelMixin
from utils.llm_utils import remove_thinking_blocks


@dataclass
class UserCode:
    """Stores analysis results for a code file"""
    file_path: str
    file_name: str
    file_content: str
    summary: str = ""
    keywords: list[str] = None
    method: str = ""
    contribution: str = ""
    important_snippets: list['CodeSnippet'] | None = None
    signatures: list[str] | None = None
    
    def __post_init__(self):
        if self.important_snippets is None:
            self.important_snippets = []
        if self.keywords is None:
            self.keywords = []
        if self.signatures is None:
            self.signatures = []


class CodeSnippet(BaseModel):
    """Represents an important code snippet extracted from a file"""
    label: str
    code: str


class UserCodeAnalysisResult(BaseModel):
    """Structured response format used by the LLM"""
    summary: str
    keywords: list[str]
    method: str
    contribution: str


class SnippetExtractionResult(BaseModel):
    """Structured response format for snippet extraction"""
    snippets: list[CodeSnippet]
    

class CodeAnalyzer(LazyModelMixin):
    """Encapsulates code file loading and LLM-based analysis methods."""

    # Supported languages
    LANGUAGE_MAP = {
        '.py': 'python',
        ".ipynb": 'jupyter notebook',
        '.js': 'javascript',
        '.java': 'java',
        '.cpp': 'cpp',
        '.c': 'c',
        '.ts': 'typescript',
        '.go': 'go',
        '.rs': 'rust',
        '.rb': 'ruby',
    }

    def __init__(self, model_name: str = "qwen/qwen3-coder-30b"):
        self.model_name = model_name
        self._model = None  # Lazy-loaded via LazyModelMixin

    @staticmethod
    def load_code_files(folder_path: str, extensions: list[str] | None = None) -> list[UserCode]:
        if extensions is None:
            extensions = list(CodeAnalyzer.LANGUAGE_MAP.keys())
        code_files = []
        folder = Path(folder_path)

        if not folder.exists():
            raise ValueError(f"Folder not found: {folder_path}")

        for file_path in folder.rglob('*'):
            if file_path.is_file() and file_path.suffix in extensions:
                try:
                    content = file_path.read_text(encoding='utf-8')
                    code_files.append(UserCode(
                        file_path=str(file_path),
                        file_name=file_path.name,
                        file_content=content
                    ))
                except Exception as e:
                    print(f"Error reading {file_path}: {e}")


        print(f"Loaded {len(code_files)} code file(s)")

        return code_files

    def analyze_code_file(self, code_analysis: UserCode) -> UserCode:
        """Analyze a code file using a single structured LLM call."""
        print(f"Analyzing {code_analysis.file_name}...")
        
        # Paper title if provided by user
        title_section = ""
        if Settings.LATEX_TITLE and Settings.LATEX_TITLE.strip():
            title_section = f"[PAPER TITLE]\n{Settings.LATEX_TITLE}\n\n"

        prompt = textwrap.dedent(f"""\
            [ROLE]
            You are a Senior Research Engineer. Extract the research essence from this code.

            [TASK]
            Analyze the code and output a JSON object.

            [GUIDELINES]
            - **Ignore Boilerplate:** Skip logging, argparse, standard IO.
            - **Be Precise:** Don't say "optimization"; say "AdamW with weight decay".

            [OUTPUT_FORMAT]
            Respond ONLY with valid JSON using exactly these keys:
            {{
                "summary": "2-3 sentences explaining what this file does technically.",
                "keywords": ["list", "of", "algorithms", "libraries", "math_terms", "architectures"],
                "method": "How is it implemented? (e.g., 'Uses a custom collate function with resizing' or 'Standard Transformer Encoder block').",
                "contribution": "What role does this play in the research? (e.g., 'Core training loop' or 'Data augmentation pipeline')."
            }}

            {title_section}[CODE FILE] 
            {code_analysis.file_name}

            [CODE CONTENT]
            ```python
            {code_analysis.file_content}
            ```"""
        )

        result = self.model.respond(
            prompt,
            response_format=UserCodeAnalysisResult,
            config={"temperature": 0.0}
        )
        # Clean thinking blocks from content before parsing
        content = remove_thinking_blocks(result.content)
        parsed_dict = json.loads(content)
        parsed = UserCodeAnalysisResult(**parsed_dict)
        code_analysis.summary = parsed.summary
        code_analysis.keywords = parsed.keywords
        code_analysis.method = parsed.method
        code_analysis.contribution = parsed.contribution

        # Debug output
        print(f"  Keywords: {', '.join(code_analysis.keywords)}")
        print(f"Completed analyzing {code_analysis.file_name}")
        return code_analysis

    def extract_important_snippets(self, code_analysis: UserCode) -> UserCode:
        """Extract important code snippets from a file."""
        
        # Paper title if provided by user
        title_section = ""
        if Settings.LATEX_TITLE and Settings.LATEX_TITLE.strip():
            title_section = f"[PAPER TITLE]\n{Settings.LATEX_TITLE}\n\n"
                
        prompt = textwrap.dedent(f"""\
            [ROLE]
            You are a Technical Editor. Your job is to select code blocks for a paper's "Methodology" section.

            {title_section}[INPUT CONTEXT]
            File: {code_analysis.file_name}
            Core Logic Identified: {code_analysis.method}
            Keywords: {', '.join(code_analysis.keywords)}

            [TASK]
            Extract 1-3 code snippets that mathematically or algorithmically implement the "Core Logic" above.

            [RULES]
            1. **Verbatim:** Copy code exactly. Do not summarize.
            2. **Pure Logic:** Exclude imports, print statements, and error handling.
            3. **Limit:** Maximum 20 lines per snippet. If it's longer, extract the central loop or equation.

            [CODE CONTENT]
            ```python
            {code_analysis.file_content}
            ```""")

        result = self.model.respond(
            prompt,
            response_format=SnippetExtractionResult,
            config={"temperature": 0.0}
        )

        content = remove_thinking_blocks(result.content)

        try:
            parsed_dict = json.loads(content)
        except json.JSONDecodeError:
            # Try to extract JSON from code fences or bare JSON
            json_match = re.search(r'```(?:json)?\s*(\{.*\})\s*```', content, re.DOTALL)
            if not json_match:
                json_match = re.search(r'(\{"snippets"\s*:\s*\[.*\].*\})', content, re.DOTALL)
            if json_match:
                try:
                    parsed_dict = json.loads(json_match.group(1))
                except json.JSONDecodeError:
                    parsed_dict = None
            else:
                parsed_dict = None

        if parsed_dict is None:
            print(f"[CodeAnalyzer] Failed to parse snippet JSON. Raw content:\n{content[:500]}")
            code_analysis.important_snippets = []
            return code_analysis

        extraction_result = SnippetExtractionResult(**parsed_dict)
        code_analysis.important_snippets = extraction_result.snippets

        print(f"Extracted {len(code_analysis.important_snippets)} code snippet(s)")
        return code_analysis

    def extract_signatures(self, code_analysis: UserCode) -> UserCode:
        """Extract function and class signatures using AST."""
        if not code_analysis.file_content:
            return code_analysis
            
        try:
            tree = ast.parse(code_analysis.file_content)
            signatures = []
            
            for node in tree.body:
                if isinstance(node, ast.FunctionDef):
                    # Extract function signature
                    args = [arg.arg for arg in node.args.args]
                    sig = f"def {node.name}({', '.join(args)})"
                    if node.returns:
                        # Simple attempt to get return type annotation as string
                        try:
                            ret_type = ast.unparse(node.returns)
                            sig += f" -> {ret_type}"
                        except Exception:
                            pass
                    signatures.append(f"Function: {sig}")
                    
                elif isinstance(node, ast.ClassDef):
                    signatures.append(f"Class: {node.name}")
                    for item in node.body:
                        if isinstance(item, ast.FunctionDef):
                            args = [arg.arg for arg in item.args.args]
                            # Only include self if it's there
                            args_str = ', '.join(args)
                            sig = f"  - method: {item.name}({args_str})"
                            if item.returns:
                                try:
                                    ret_type = ast.unparse(item.returns)
                                    sig += f" -> {ret_type}"
                                except Exception:
                                    pass
                            signatures.append(sig)
            
            code_analysis.signatures = signatures
            print(f"Extracted {len(signatures)} signatures from {code_analysis.file_name}")
            
        except Exception as e:
            print(f"Error parsing AST for {code_analysis.file_name}: {e}")
            
        return code_analysis

    def analyze_all_files(self, code_files: list[UserCode], extract_signatures: bool = True) -> list[UserCode]:
        """
        Analyze all code files and extract important code snippets.
        """
        analyzed_files = []
        for code_file in code_files:
            analyzed = self.analyze_code_file(code_file)
            analyzed = self.extract_important_snippets(analyzed)
            
            if extract_signatures:
                analyzed = self.extract_signatures(analyzed)
                
            analyzed_files.append(analyzed)
        
        print(f"Code analysis complete: analyzed {len(analyzed_files)} file(s)")
        return analyzed_files
    

    @staticmethod
    def get_analysis_report(analyzed_files: list[UserCode]) -> str:
        report = []
        
        for analysis in analyzed_files:
            report.extend([
                f"## File: {analysis.file_name}",
                f"\n**Summary:** {analysis.summary}"
            ])
            
            if analysis.keywords:
                report.append(f"\n**Keywords:** {', '.join(analysis.keywords)}")
            
            if analysis.method:
                report.append(f"\n**Method:** {analysis.method}")
                
            if analysis.contribution:
                report.append(f"\n**Contribution:** {analysis.contribution}")
            
            if analysis.important_snippets:
                report.append(f"\n**Code Snippets ({len(analysis.important_snippets)}):**")
                for snippet in analysis.important_snippets:
                    report.extend([
                        f"\n### {snippet.label}",
                        f"```python\n{snippet.code}\n```"
                    ])
            
            report.append("\n---\n")

        return "\n".join(report)


# ==================== Dataset Analysis ====================

DATASET_EXTENSIONS = {'.csv', '.tsv', '.json', '.jsonl', '.xlsx', '.xls', '.parquet'}


@dataclass
class UserDataset:
    """Stores metadata and analysis results for a dataset file."""
    file_path: str
    file_name: str
    file_size: int  # bytes
    columns: list[str] = field(default_factory=list)
    row_count: int = 0
    dtypes: str = ""
    raw_head: str = ""  # first few raw lines as fallback preview
    load_instruction: str = ""

    @property
    def file_size_display(self) -> str:
        """Human-readable file size."""
        if self.file_size < 1024:
            return f"{self.file_size} B"
        elif self.file_size < 1024 * 1024:
            return f"{self.file_size / 1024:.1f} KB"
        else:
            return f"{self.file_size / (1024 * 1024):.1f} MB"


class DatasetAnalyzer:
    """Analyzes dataset files to extract metadata without loading full content."""

    LOAD_INSTRUCTIONS = {
        '.csv': 'pd.read_csv("{path}")',
        '.tsv': 'pd.read_csv("{path}", sep="\\t")',
        '.json': 'pd.read_json("{path}")',
        '.jsonl': 'pd.read_json("{path}", lines=True)',
        '.xlsx': 'pd.read_excel("{path}")',
        '.xls': 'pd.read_excel("{path}")',
        '.parquet': 'pd.read_parquet("{path}")',
    }

    @staticmethod
    def load_datasets(folder_path: str) -> list[UserDataset]:
        """Load dataset file metadata from a folder."""
        folder = Path(folder_path)
        datasets = []

        if not folder.exists():
            return datasets

        for file_path in folder.iterdir():
            if file_path.is_file() and file_path.suffix.lower() in DATASET_EXTENSIONS:
                datasets.append(UserDataset(
                    file_path=str(file_path),
                    file_name=file_path.name,
                    file_size=file_path.stat().st_size,
                ))

        print(f"Found {len(datasets)} dataset file(s)")
        return datasets

    @staticmethod
    def _read_raw_head(file_path: Path, n_lines: int = 4) -> str:
        """Read the first n lines of a text file as raw strings."""
        try:
            with open(file_path, 'r', encoding='utf-8-sig', errors='replace') as f:
                lines = [next(f).rstrip('\n\r') for _ in range(n_lines)]
            return "\n".join(lines)
        except StopIteration:
            # File has fewer lines than requested
            with open(file_path, 'r', encoding='utf-8-sig', errors='replace') as f:
                return f.read().rstrip()
        except Exception:
            return ""

    @staticmethod
    def analyze_dataset(dataset: UserDataset) -> UserDataset:
        """Analyze a dataset file to extract columns, dtypes, row count, and sample rows."""
        file_path = Path(dataset.file_path)
        suffix = file_path.suffix.lower()

        # Always capture raw head for text-based formats
        if suffix not in ('.parquet', '.xlsx', '.xls'):
            dataset.raw_head = DatasetAnalyzer._read_raw_head(file_path)

        try:
            import pandas as pd
        except ImportError:
            print("Warning: pandas not available, skipping dataset analysis")
            return dataset

        try:
            # Read a small sample for schema + preview
            if suffix == '.csv':
                df = pd.read_csv(file_path, nrows=5, sep=None, engine='python', encoding='utf-8-sig', encoding_errors='replace')
            elif suffix == '.tsv':
                df = pd.read_csv(file_path, sep='\t', nrows=5, encoding_errors='replace')
            elif suffix == '.json':
                df = pd.read_json(file_path)
                df = df.head(5)
            elif suffix == '.jsonl':
                df = pd.read_json(file_path, lines=True, nrows=5)
            elif suffix in ('.xlsx', '.xls'):
                df = pd.read_excel(file_path, nrows=5)
            elif suffix == '.parquet':
                try:
                    import pyarrow.parquet as pq
                    parquet_file = pq.ParquetFile(file_path)
                    full_len = parquet_file.metadata.num_rows
                    df = parquet_file.read().to_pandas().head(5)
                    dataset.row_count = full_len
                except (ImportError, Exception):
                    try:
                        df = pd.read_parquet(file_path)
                        dataset.row_count = len(df)
                        df = df.head(5)
                    except ImportError:
                        print(f"Warning: No parquet engine available, skipping {dataset.file_name}")
                        return dataset
            else:
                return dataset

            # Get row count efficiently (skip for parquet, already handled)
            if suffix != '.parquet':
                if suffix in ('.csv', '.tsv'):
                    sep = '\t' if suffix == '.tsv' else None
                    engine = 'python' if sep is None else 'c'
                    full_len = sum(len(chunk) for chunk in pd.read_csv(
                        file_path, sep=sep, engine=engine, chunksize=10000, usecols=[0],
                        encoding='utf-8-sig', encoding_errors='replace'
                    ))
                elif suffix == '.json':
                    full_len = len(pd.read_json(file_path))
                elif suffix == '.jsonl':
                    full_len = sum(len(chunk) for chunk in pd.read_json(file_path, lines=True, chunksize=10000))
                elif suffix in ('.xlsx', '.xls'):
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(file_path, read_only=True)
                        ws = wb.active
                        full_len = ws.max_row - 1 if ws.max_row else 0
                        wb.close()
                    except Exception:
                        full_len = len(pd.read_excel(file_path))
                dataset.row_count = full_len

            dataset.columns = list(df.columns)
            dataset.dtypes = ", ".join(f"{col} ({dtype})" for col, dtype in zip(df.columns, df.dtypes))

            # Build load instruction with relative path for experiments
            rel_path = f"datasets/{dataset.file_name}"
            template = DatasetAnalyzer.LOAD_INSTRUCTIONS.get(suffix, 'pd.read_csv("{path}")')
            dataset.load_instruction = template.format(path=rel_path)

        except Exception as e:
            print(f"Error analyzing dataset {dataset.file_name}: {e}")

        return dataset

    @staticmethod
    def analyze_all_datasets(datasets: list[UserDataset]) -> list[UserDataset]:
        """Analyze all dataset files."""
        analyzed = []
        for ds in datasets:
            analyzed.append(DatasetAnalyzer.analyze_dataset(ds))
        print(f"Dataset analysis complete: analyzed {len(analyzed)} file(s)")
        return analyzed

    @staticmethod
    def get_dataset_report(datasets: list[UserDataset]) -> str:
        """Format dataset metadata into a readable report."""
        if not datasets:
            return ""

        report = []

        for ds in datasets:
            col_count = len(ds.columns)
            report.append(f"## {ds.file_name}")
            report.append(f"**Size:** {ds.file_size_display} | **Rows:** {ds.row_count:,} | **Columns:** {col_count}")

            if ds.dtypes:
                report.append(f"\n**Column types:** {ds.dtypes}")

            if ds.raw_head:
                report.append(f"\n**Raw preview:**\n```\n{ds.raw_head}\n```")

            report.append("\n---\n")

        return "\n".join(report)
